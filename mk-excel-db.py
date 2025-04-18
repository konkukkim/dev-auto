# for git hub dist
import pymysql
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.styles.colors import Color
from datetime import datetime
import sys

now = datetime.now().strftime('%Y%m%d%H%M%S')

# 테이블 이름 목록을 커맨드라인 인자에서 받아오기
table_names = sys.argv[1:]

doc_writer = ""

doc_now = datetime.now().strftime('%Y-%m-%d')

# MariaDB 연결 정보
host = "123.45.67.89"
user = "iamuser"
password = "iampass"
database = "iamdb"

# MariaDB 연결
connection = pymysql.connect(
        host=host,    # MariaDB 호스트
        user=user,         # MariaDB 사용자
        password=password, # MariaDB 비밀번호
        database=database    # 사용할 데이터베이스
        )

# 테이블 설명 가져오기
def get_table_comment(table_name):
    with connection.cursor() as cursor:
        cursor.execute(f""" SELECT 
                        TABLE_COMMENT
                    FROM 
                        INFORMATION_SCHEMA.TABLES
                    WHERE 
                        TABLE_NAME = '{table_name}' """)
        comment = cursor.fetchone()  
        return comment


# 테이블 스키마 가져오기
def get_table_schema(table_name):
    with connection.cursor() as cursor:
        cursor.execute(f""" SELECT ORDINAL_POSITION AS 번호
	,'' as 'AS-IS 컬럼명(물리)'
    ,column_name as '컬럼명(물리)'
    ,column_comment  as '속성명(논리)'
    , '' as '도메인'
    , upper(column_type) as '데이터타입'
    ,case when is_nullable = 'YES' then 'Y' else 'N' end as 'NULL허용'
    , COLUMN_DEFAULT
    , case when COLUMN_key = 'PRI' then 'PK' 
           when COLUMN_key = 'UNI' then 'UK'
           else COLUMN_key
      end as 'KEY'
     , '' as 코멘트
FROM
    information_schema.columns
where 1=1 
 and table_name = '{table_name}' """)
        schema = cursor.fetchall()  # 테이블의 컬럼 정보 가져오기
        return schema

def write_header(data, wb, ws, start_row):

    # 폰트 설정
    font_style = Font(name="Malgun Gothic", size=10)

    # 데이터 작성 및 폰트 적용
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, cell_data in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_data)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = font_style  # 폰트 적용

    # 셀 병합 및 외부 윤곽선 추가
    ws.merge_cells(f"A{start_row}:B{start_row}")  # 셀 병합
    ws.merge_cells(f"A{start_row+1}:B{start_row+1}")  # 셀 병합
    ws.merge_cells(f"A{start_row+2}:B{start_row+2}")  # 셀 병합
    ws.merge_cells(f"C{start_row+2}:H{start_row+2}")  # 셀 병합
    ws.merge_cells(f"D{start_row}:E{start_row}")  # 셀 병합
    ws.merge_cells(f"D{start_row+1}:E{start_row+1}")  # 셀 병합
    ws.merge_cells(f"F{start_row}:H{start_row}")  # 셀 병합
    ws.merge_cells(f"F{start_row+1}:H{start_row+1}")  # 셀 병합
#    ws.merge_cells(f"I{start_row}:I{start_row+2}")  # 셀 병합
#    ws.merge_cells(f"J{start_row}:J{start_row+2}")  # 셀 병합
#    ws.cell(row=start_row, column=9).value = "코멘트"  # 병합된 셀에 "코멘트" 입력
#    ws.cell(row=start_row, column=9).alignment = Alignment(horizontal='center', vertical='center') # 가운데 정렬
#    ws.cell(row=start_row, column=9).font = font_style  # 폰트 적용

    # 외부 윤곽선 스타일 정의
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # 병합된 셀에 외부 윤곽선 적용
    # for row in ws[f"I{start_row}:I{start_row+2}"]:
    #     for cell in row:
    #         cell.border = border_style
    # for row in ws[f"J{start_row}:J{start_row+2}"]:
    #     for cell in row:
    #         cell.border = border_style

    # 나머지 셀에 외부 윤곽선 적용
    for row in ws[f"A{start_row}:H{start_row+2}"]:
        for cell in row:
            cell.border = border_style

    # 회색 음영 적용
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 회색 음영 스타일 정의
    for row in ws[f"A{start_row}:B{start_row+2}"]:
        for cell in row:
            cell.fill = gray_fill  # 회색 음영 적용
    for row in ws[f"D{start_row}:E{start_row+1}"]:
        for cell in row:
            cell.fill = gray_fill  # 회색 음영 적용
    # for row in ws[f"I{start_row}:I{start_row+2}"]:
    #     for cell in row:
    #         cell.fill = gray_fill  # 회색 음영 적용

    return wb


# 엑셀 문서 생성
def write_schema(schema, wb, ws, start_row):

    # 헤더 추가
    ws.append(["번호", "AS-IS 컬럼명(물리) ", "컬럼명(물리)", "속성명(논리)", "도메인", "데이터타입", "NULL 허용", "기본값", "KEY", "코멘트"])

    # 테이블 스키마 추가
    for row in schema:
        ws.append(row)

    # 외부 윤곽선 스타일 정의
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # 첫 번째 줄에 적용할 폰트 (굵은 글씨)
    bold_font = Font(name="Malgun Gothic", size=10, bold=True)


    # 셀에 외부 윤곽선 적용
    # last_row_no = start_row + 5 + len(schema)
    last_row_no = start_row + 4 + len(schema)
    cell_end = f"J{last_row_no}"
    for row in ws[f"A{start_row+5}:{cell_end}"]:
        for cell in row:
            cell.border = border_style

    print (f"start_row:{start_row}")
    print (f"last_row_no:{last_row_no}")


    # 헤더 부분에 회색 음영 적용
    gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')  # 회색 음영 스타일 정의
    # for row in ws[f"A{start_row+4}:J{start_row+4}"]:  # 셀 선택
    for row in ws[f"A{start_row}:B{start_row+2}"]:  # 셀 선택
        for cell in row:
            cell.fill = gray_fill  # 회색 음영 적용
            cell.border = border_style
            cell.font = bold_font
    for row in ws[f"D{start_row}:E{start_row+1}"]:  # 셀 선택
        for cell in row:
            cell.fill = gray_fill  # 회색 음영 적용
            cell.border = border_style
            cell.font = bold_font
    # for row in ws[f"I{start_row}:I{start_row+1}"]:  # 셀 선택
    #     for cell in row:
    #         cell.fill = gray_fill  # 회색 음영 적용
    #         cell.border = border_style
    #         cell.font = bold_font
    for row in ws[f"A{start_row+4}:J{start_row+4}"]:  # 셀 선택
        for cell in row:
            cell.fill = gray_fill  # 회색 음영 적용
            cell.border = border_style
            cell.font = bold_font
    
    # 폰트 설정: 기본 폰트 (Malgun Gothic, 크기 10)
    default_font = Font(name="Malgun Gothic", size=10)
    
    print (f"start_row+5:{start_row+5}")

    # 워크시트의 모든 셀에 기본 폰트 적용
    for row in ws[f"A{start_row+5}:J{last_row_no}"]:
        for cell in row:
            cell.font = default_font


    return wb


# 메인 실행
if __name__ == "__main__":

    wb = Workbook()
    ws = wb.active
    ws.title = "Table Schema"

    # 테이블들을 순차적으로 처리하고 엑셀에 작성
    start_row = 1
    for table_name in table_names:
        comment = get_table_comment(table_name)
        if comment is None:
            print(f"{table_name} 스키마가 없습니다")
            continue

        # 각 테이블에 대한 데이터 준비
        table_data = [
            ["엔티티 타입 명 (논리)",	"", f"{comment[0]}", 	"작성일", "",	f"{doc_now}",     "", ""],
            ["테이블 명 (물리)", 		"", f"{table_name}",	"작성자", "",	f"{doc_writer}", "", ""],
            ["테이블 설명", 			"", f"{comment[0]}",	"", 	 "",  	"",			    "", ""],
            ["", "", "","","","","","", "",""],
        ]

        # 각 테이블의 명세 작성
        wb = write_header(table_data, wb, ws, start_row)
        schema = get_table_schema(table_name)
        wb = write_schema(schema, wb, ws, start_row)

        # 테이블 간 간격을 주기 위해 start_row 갱신
        start_row += 5 + len(schema) + 1  # 간격을 10으로 설정

    # 엑셀 파일 저장
    wb.save(f"{now}.xlsx")
    print(f"{now}.xlsx")

    connection.close()  # 연결 종료
